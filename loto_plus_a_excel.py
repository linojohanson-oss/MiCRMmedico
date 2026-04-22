import os
import random
import re
import sys
import time
from dataclasses import dataclass
from typing import List, Optional, Tuple, Set, Dict

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# =========================
# FUENTE
# =========================
HOME_URL = "https://soytimbero.com/lotoplus"
DRAW_URL_FMT = "https://soytimbero.com/lotoplus/sorteo-{}"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0 Safari/537.36"
)

# Pausa recomendada (para 500 o incremental)
PAUSE_MIN = 1.5
PAUSE_MAX = 3.0

# =========================
# MODELO
# =========================
@dataclass
class SorteoLotoPlus:
    sorteo: int
    fecha: str
    numero_plus: Optional[int]
    tradicional: List[int]
    match: List[int]
    desquite: List[int]
    sale_o_sale: List[int]


# =========================
# HTTP ROBUSTO
# =========================
_session: Optional[requests.Session] = None


def _make_session() -> requests.Session:
    global _session
    if _session is not None:
        return _session

    s = requests.Session()
    retries = Retry(
        total=6,
        connect=6,
        read=6,
        backoff_factor=1.2,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retries, pool_connections=10, pool_maxsize=10)
    s.mount("https://", adapter)
    s.mount("http://", adapter)

    s.headers.update({
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-AR,es;q=0.9,en;q=0.7",
        "Connection": "keep-alive",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Upgrade-Insecure-Requests": "1",
    })

    _session = s
    return s


def _get(url: str, timeout: int = 25, human_pause: bool = True) -> str:
    s = _make_session()
    if human_pause:
        time.sleep(random.uniform(PAUSE_MIN, PAUSE_MAX))

    r = s.get(url, timeout=timeout)
    if r.status_code in (403, 429):
        time.sleep(random.uniform(4, 9))
        r = s.get(url, timeout=timeout)

    r.raise_for_status()
    return r.text


# =========================
# PARSE
# =========================
def _normalize_text(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text


def _extract_6_after_label(text: str, label: str) -> List[int]:
    i = text.lower().find(label.lower())
    if i == -1:
        return []
    chunk = text[i:i + 500]
    nums = re.findall(r"\b\d{1,2}\b", chunk)
    out = [int(n) for n in nums]
    return out[:6] if len(out) >= 6 else []


def _extract_numero_plus(text: str) -> Optional[int]:
    m = re.search(r"N[uú]mero Plus\s*\n\s*([0-9]{1,2})\b", text, flags=re.IGNORECASE)
    if not m:
        m = re.search(r"N[uú]mero Plus\s*([0-9]{1,2})\b", text, flags=re.IGNORECASE)
    if not m:
        return None
    v = int(m.group(1))
    return v if 0 <= v <= 9 else None


def _extract_header(text: str) -> Tuple[int, str]:
    m = re.search(r"Sorteo\s+(\d+)", text, flags=re.IGNORECASE)
    if not m:
        raise ValueError("No pude encontrar el número de sorteo.")
    sorteo = int(m.group(1))

    lines = text.splitlines()
    fecha = ""
    for ln in lines[:40]:
        if re.search(r"\bde\b.*\b20\d{2}\b", ln, flags=re.IGNORECASE):
            fecha = ln.strip()
            break
    if not fecha:
        for ln in lines[:60]:
            if re.search(r"\b20\d{2}\b", ln):
                fecha = ln.strip()
                break

    return sorteo, fecha


def parse_sorteo(html: str) -> SorteoLotoPlus:
    text = _normalize_text(html)
    sorteo, fecha = _extract_header(text)
    numero_plus = _extract_numero_plus(text)

    tradicional = _extract_6_after_label(text, "Tradicional")
    match_nums = _extract_6_after_label(text, "Match")
    desquite = _extract_6_after_label(text, "Desquite")
    sale_o_sale = _extract_6_after_label(text, "Sale o Sale")

    if not (tradicional and match_nums and desquite and sale_o_sale):
        raise ValueError(f"Sorteo {sorteo}: faltan números en alguna modalidad (cambió el formato).")

    return SorteoLotoPlus(
        sorteo=sorteo,
        fecha=fecha,
        numero_plus=numero_plus,
        tradicional=tradicional,
        match=match_nums,
        desquite=desquite,
        sale_o_sale=sale_o_sale,
    )


def get_latest_draw_number() -> int:
    html = _get(HOME_URL, human_pause=False)
    text = _normalize_text(html)
    m = re.search(r"Sorteo\s+(\d+)", text, flags=re.IGNORECASE)
    if not m:
        raise RuntimeError("No pude detectar el último número de sorteo en la página principal.")
    return int(m.group(1))


# =========================
# LECTURA / MERGE INCREMENTAL
# =========================
SHEETS = ["Resumen", "Tradicional", "Match", "Desquite", "Sale o Sale"]


def read_existing_excel(path: str) -> Dict[str, pd.DataFrame]:
    """
    Lee las hojas si existen. Si no existe el archivo, devuelve dict vacío.
    """
    if not os.path.exists(path):
        return {}
    dfs = {}
    xls = pd.ExcelFile(path)
    for s in SHEETS:
        if s in xls.sheet_names:
            dfs[s] = pd.read_excel(path, sheet_name=s)
    return dfs


def existing_sorteos(existing: Dict[str, pd.DataFrame]) -> Set[int]:
    """
    Obtiene set de sorteos ya presentes (usa 'Resumen' si está).
    """
    if "Resumen" in existing and not existing["Resumen"].empty and "sorteo" in existing["Resumen"].columns:
        return set(existing["Resumen"]["sorteo"].astype(int).tolist())

    # fallback: intentar otra hoja
    for k, df in existing.items():
        if "sorteo" in df.columns and not df.empty:
            return set(df["sorteo"].astype(int).tolist())

    return set()


# =========================
# DATAFRAMES / EXCEL
# =========================
def _df_modalidad(items: List[SorteoLotoPlus], key: str) -> pd.DataFrame:
    rows = []
    for it in items:
        nums = getattr(it, key)
        row = {"sorteo": it.sorteo, "fecha": it.fecha}
        if key == "tradicional":
            row["numero_plus"] = it.numero_plus
        for i, n in enumerate(nums, 1):
            row[f"n{i}"] = n
        rows.append(row)
    return pd.DataFrame(rows)


def _df_resumen(items: List[SorteoLotoPlus]) -> pd.DataFrame:
    rows = []
    for it in items:
        rows.append({
            "sorteo": it.sorteo,
            "fecha": it.fecha,
            "numero_plus": it.numero_plus,
            "tradicional": "-".join(f"{n:02d}" for n in it.tradicional),
            "match": "-".join(f"{n:02d}" for n in it.match),
            "desquite": "-".join(f"{n:02d}" for n in it.desquite),
            "sale_o_sale": "-".join(f"{n:02d}" for n in it.sale_o_sale),
        })
    return pd.DataFrame(rows)


def _style_sheet(ws, df: pd.DataFrame):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), *(len(str(v)) for v in df[col_name].astype(str).values))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 36)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center


def export_excel(dfs: Dict[str, pd.DataFrame], path: str) -> None:
    """
    Escribe todas las hojas desde el dict dfs.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name in SHEETS:
            df = dfs.get(name, pd.DataFrame())
            df.to_excel(writer, sheet_name=name, index=False)

        for name in SHEETS:
            df = dfs.get(name, pd.DataFrame())
            ws = writer.sheets[name]
            if not df.empty:
                _style_sheet(ws, df)


def merge_and_sort(existing: Dict[str, pd.DataFrame], new_items: List[SorteoLotoPlus]) -> Dict[str, pd.DataFrame]:
    new_resumen = _df_resumen(new_items)
    new_trad = _df_modalidad(new_items, "tradicional")
    new_match = _df_modalidad(new_items, "match")
    new_desq = _df_modalidad(new_items, "desquite")
    new_sale = _df_modalidad(new_items, "sale_o_sale")

    def _merge(sheet_name: str, new_df: pd.DataFrame) -> pd.DataFrame:
        old_df = existing.get(sheet_name, pd.DataFrame())
        if old_df is None or old_df.empty:
            merged = new_df.copy()
        else:
            merged = pd.concat([old_df, new_df], ignore_index=True)

        if "sorteo" in merged.columns:
            merged["sorteo"] = merged["sorteo"].astype(int)
            merged = merged.drop_duplicates(subset=["sorteo"], keep="first")
            merged = merged.sort_values("sorteo", ascending=False).reset_index(drop=True)
        return merged

    out = {}
    out["Resumen"] = _merge("Resumen", new_resumen)
    out["Tradicional"] = _merge("Tradicional", new_trad)
    out["Match"] = _merge("Match", new_match)
    out["Desquite"] = _merge("Desquite", new_desq)
    out["Sale o Sale"] = _merge("Sale o Sale", new_sale)
    return out


# =========================
# MAIN
# =========================
def main(n_ultimos: int = 500, out_xlsx: str = "loto_plus_incremental.xlsx") -> None:
    existing = read_existing_excel(out_xlsx)
    ya_tengo = existing_sorteos(existing)

    latest = get_latest_draw_number()

    # Querés tener al menos los últimos n_ultimos en el Excel.
    # Si ya hay datos, igual revisamos desde latest hacia atrás hasta completar faltantes.
    target_draws = list(range(latest, latest - n_ultimos, -1))

    to_fetch = [d for d in target_draws if d not in ya_tengo]

    if not to_fetch and os.path.exists(out_xlsx):
        print(f"No hay sorteos nuevos. El Excel ya tiene los últimos {n_ultimos} (o más).")
        print(f"Archivo: {os.path.abspath(out_xlsx)}")
        return

    print(f"Excel actual: {len(ya_tengo)} sorteos. Faltan descargar: {len(to_fetch)}")

    items: List[SorteoLotoPlus] = []
    failed: List[int] = []

    total = len(to_fetch)
    for idx, d in enumerate(to_fetch, start=1):
        try:
            html = _get(DRAW_URL_FMT.format(d), human_pause=True)
            it = parse_sorteo(html)
            items.append(it)
            print(f"✓ Sorteo {d} OK ({idx}/{total})")
        except Exception as e:
            failed.append(d)
            print(f"✗ Sorteo {d} FALLÓ ({idx}/{total}): {e}")

    if not items and not os.path.exists(out_xlsx):
        raise RuntimeError("No se pudo procesar ningún sorteo. Posible bloqueo o cambio fuerte de formato.")

    merged = merge_and_sort(existing, items) if items else existing
    export_excel(merged, out_xlsx)

    print(f"\nOK ✅ Guardado/actualizado: {os.path.abspath(out_xlsx)}")
    print(f"Sorteos nuevos agregados: {len(items)}")
    if failed:
        print(f"Sorteos fallidos (para reintentar): {failed}")


if __name__ == "__main__":
    # Uso:
    #   python loto_plus_incremental.py
    #   python loto_plus_incremental.py 500 loto_plus_incremental.xlsx
    #   python loto_plus_incremental.py 1000 loto_plus_incremental.xlsx
    n = int(sys.argv[1]) if len(sys.argv) >= 2 else 500
    out = sys.argv[2] if len(sys.argv) >= 3 else "loto_plus_incremental.xlsx"
    main(n_ultimos=n, out_xlsx=out)
