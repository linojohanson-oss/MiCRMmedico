from flask import Blueprint, render_template, send_file
from db import obtener_conexion
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

reportes_bp = Blueprint('reportes', __name__)


@reportes_bp.route('/reportes')
def ver_reportes():
    conn = obtener_conexion(dict_cursor=True)
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM clientes")
    total_clientes = cursor.fetchone()['total']

    cursor.execute("SELECT COUNT(*) AS total FROM doctores")
    total_doctores = cursor.fetchone()['total']

    cursor.execute("SELECT COUNT(*) AS total FROM citas")
    total_citas = cursor.fetchone()['total']

    cursor.execute("SELECT COUNT(*) AS total FROM atenciones")
    total_atenciones = cursor.fetchone()['total']

    cursor.execute("""
        SELECT d.especialidad, COUNT(*) AS total
        FROM atenciones a
        JOIN citas c ON a.cita_id = c.id
        JOIN doctores d ON c.doctor_id = d.id
        GROUP BY d.especialidad
        ORDER BY total DESC
    """)
    atenciones_por_especialidad = cursor.fetchall()

    cursor.execute("""
        SELECT DATE_FORMAT(fecha, '%Y-%m') AS mes, COUNT(*) AS total
        FROM citas
        GROUP BY mes
        ORDER BY mes
    """)
    citas_por_mes = cursor.fetchall()

    cursor.execute("""
        SELECT d.nombre AS doctor, COUNT(a.id) AS total
        FROM atenciones a
        JOIN citas c ON a.cita_id = c.id
        JOIN doctores d ON c.doctor_id = d.id
        GROUP BY d.nombre
        ORDER BY total DESC
    """)
    atenciones_por_doctor = cursor.fetchall()

    cursor.execute("""
        SELECT cl.nombre AS paciente, COUNT(c.id) AS total
        FROM citas c
        JOIN clientes cl ON c.cliente_id = cl.id
        GROUP BY cl.nombre
        ORDER BY total DESC
        LIMIT 5
    """)
    pacientes_frecuentes = cursor.fetchall()

    cursor.execute("""
        SELECT d.nombre AS doctor, COUNT(c.id) AS total
        FROM citas c
        JOIN doctores d ON c.doctor_id = d.id
        GROUP BY d.nombre
        ORDER BY total DESC
    """)
    citas_por_doctor = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'reportes.html',
        total_clientes=total_clientes,
        total_doctores=total_doctores,
        total_citas=total_citas,
        total_atenciones=total_atenciones,
        atenciones_por_especialidad=atenciones_por_especialidad,
        citas_por_mes=citas_por_mes,
        atenciones_por_doctor=atenciones_por_doctor,
        pacientes_frecuentes=pacientes_frecuentes,
        citas_por_doctor=citas_por_doctor
    )


@reportes_bp.route('/reportes/exportar_excel')
def exportar_excel():
    conn = obtener_conexion(dict_cursor=True)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()

    cursor.execute("SELECT * FROM doctores")
    doctores = cursor.fetchall()

    cursor.execute("SELECT * FROM citas")
    citas = cursor.fetchall()

    cursor.execute("SELECT * FROM atenciones")
    atenciones = cursor.fetchall()

    cursor.execute("""
        SELECT d.especialidad, COUNT(*) AS total
        FROM atenciones a
        JOIN citas c ON a.cita_id = c.id
        JOIN doctores d ON c.doctor_id = d.id
        GROUP BY d.especialidad
        ORDER BY total DESC
    """)
    resumen_especialidades = cursor.fetchall()

    cursor.execute("""
        SELECT d.nombre AS doctor, COUNT(a.id) AS total
        FROM atenciones a
        JOIN citas c ON a.cita_id = c.id
        JOIN doctores d ON c.doctor_id = d.id
        GROUP BY d.nombre
        ORDER BY total DESC
    """)
    resumen_doctores = cursor.fetchall()

    cursor.execute("""
        SELECT cl.nombre AS paciente, COUNT(c.id) AS total
        FROM citas c
        JOIN clientes cl ON c.cliente_id = cl.id
        GROUP BY cl.nombre
        ORDER BY total DESC
    """)
    resumen_pacientes = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()

    crear_hoja(wb.active, "Clientes", clientes)
    crear_hoja(wb.create_sheet("Doctores"), "Doctores", doctores)
    crear_hoja(wb.create_sheet("Citas"), "Citas", citas)
    crear_hoja(wb.create_sheet("Atenciones"), "Atenciones", atenciones)
    crear_hoja(wb.create_sheet("Resumen Especialidades"), "Resumen Especialidades", resumen_especialidades)
    crear_hoja(wb.create_sheet("Resumen Doctores"), "Resumen Doctores", resumen_doctores)
    crear_hoja(wb.create_sheet("Resumen Pacientes"), "Resumen Pacientes", resumen_pacientes)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_crm_medico.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def crear_hoja(ws, titulo, datos):
    ws.title = titulo

    ws["A1"] = f"Reporte de {titulo}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2C3E50")
    ws["A1"].alignment = Alignment(horizontal="center")

    if not datos:
        ws["A3"] = "No hay datos disponibles."
        return

    columnas = list(datos[0].keys())

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(columnas)
    )

    for col_num, columna in enumerate(columnas, 1):
        celda = ws.cell(row=3, column=col_num)
        celda.value = columna
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="3498DB")
        celda.alignment = Alignment(horizontal="center")

    for row_num, fila in enumerate(datos, 4):
        for col_num, columna in enumerate(columnas, 1):
            ws.cell(row=row_num, column=col_num).value = fila[columna]

    for col_num in range(1, len(columnas) + 1):
        letra = get_column_letter(col_num)
        ws.column_dimensions[letra].width = 22